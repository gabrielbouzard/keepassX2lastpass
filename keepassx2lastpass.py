#!/usr/bin/python
#title			:keepassx2lastpass.py
#description		:convert keepassx CVS excel format to lastpass for easy import
#author			:gabriel bouzard
#date			:20170820
#version		:0.1
#usage			:python keepassx2lastpass.py
#notes			:URL, Name, Folder, Username, Password, Notes
#python_version		:3.5
#==============================================================================

import argparse
import openpyxl

def main():

    parser = argparse.ArgumentParser()
    parser.add_argument('-i', '--in', help="KeepassX CVS Excel file", required=True)
    parser.add_argument('-o', '--out', help="LastPass import file", required=True)
    parser.add_argument('-s', '--system', help="W for Windows, L for Linux, A for OSX")
    args = vars(parser.parse_args())

    newline = ""

    kp = openpyxl.load_workbook(args['in'])
    lp = open("last_pass_cvs.txt", "w")

    if (args["system"] == "W"):
        newline = "\r\n"
    elif (args["system"] == "L"):
        newline = "\n"
    elif (args["system"] == "A"):
        newline = "\r"
    else:
        newline = "\n"
    
    for sheet in kp.worksheets:  

        lp.write("url,username,password,extra,name,grouping,fav" + newline)

        for r in range(1, sheet.max_row + 1):
            group = str(sheet.cell(row=r, column=1).value)
            title = str(sheet.cell(row=r, column=2).value)
            username = str(sheet.cell(row=r, column=3).value)
            password = str(sheet.cell(row=r, column=4).value)
            url = str(sheet.cell(row=r, column=5).value)
            notes = str(sheet.cell(row=r, column=6).value)

            lp.write("http://" + url + "," + username + "," + password + "," + group + " ::: " + notes.strip() + "," + title + ",,0" + newline)

    lp.close()
            
if __name__ == '__main__':
    main()
